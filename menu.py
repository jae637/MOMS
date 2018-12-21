import urllib.request
import datetime

from bs4 import BeautifulSoup
from openpyxl import load_workbook
#10~,5~9

def load(name,date,a):
    file_name=name+'.xlsx'
    wb=load_workbook(file_name)
    sheet1 = wb['Sheet1']
    #sheet2= wb.active
    if date== 1:
        datestr ='E'
    elif date ==2:
        datestr ='F'
    elif date ==3:
        datestr ='G'
    elif date ==4:
        datestr ='H'
    elif date ==5:
        datestr ='I'

    menu = []

    for i in range(a,a+7):
        var = datestr+str(i)
        menu.append(sheet1[var].value)

    menu.append("\n총 칼로리는:question-box:  "+sheet1[datestr+str(a+7)].value+":bomb:")

    return menu

def recommend(name,date):
    file_name=name+'.xlsx'
    wb=load_workbook(file_name)
    sheet1 = wb['Sheet1']

    mes=[]

    if date== 1:
        datestr ='E'
    elif date ==2:
        datestr ='F'
    elif date ==3:
        datestr ='G'
    elif date ==4:
        datestr ='H'
    elif date ==5:
        datestr ='I'

    if(int(sheet1[datestr+str(12)].value.split()[0])>int(sheet1[datestr+str(20)].value.split()[0])):
        #5
        mes=load(name,date,5)
        mes.append('A식단이 칼로리가 더 높아서 맛있습니다.')
    else :
        #13
        mes=load(name,date,13)
        mes.append('B식단이 칼로리가 더 높아서 맛있습니다.')

    return mes
